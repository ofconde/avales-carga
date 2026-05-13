#!/usr/bin/env python3
"""
Importa todos los registros del Excel (hoja 'trasladable') a la DB local.
Uso:  python3 migrate_excel.py
      python3 migrate_excel.py /ruta/al/archivo.xlsx
"""
import sys
import json
import sqlite3
import re
from pathlib import Path
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    print("ERROR: Instalar openpyxl primero →  pip3 install openpyxl")
    sys.exit(1)

BASE_DIR = Path(__file__).parent
DB_PATH  = BASE_DIR / 'carga.db'

EXCEL_DEFAULT = Path.home() / "Library/CloudStorage/OneDrive-SharedLibraries-CONSEJOFEDERALDEINVERSIONES/Carga e Instrumentación de Créditos - PLANILLAS VARIAS DE SOLICITUD CREDITOS/CARGA DE EXPEDIENTES 2025.xlsx"

def fmt_date(val):
    if val is None:
        return None
    if isinstance(val, (datetime, date)):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s or s.lower() in ('nan', 'none', '-', ''):
        return None
    # DD/MM/YYYY
    m = re.match(r'^(\d{1,2})[/\-](\d{1,2})[/\-](\d{2,4})$', s)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2:
            y = '20' + y
        return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
    # YYYY-MM-DD
    m = re.match(r'^(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})$', s)
    if m:
        y, mo, d = m.groups()
        return f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
    return None

def clean_monto(val):
    if val is None:
        return ''
    s = str(val).strip()
    if s.lower() in ('nan', 'none', '-', ''):
        return ''
    # Eliminar $ y espacios, normalizar separadores argentinos
    s = s.replace('$', '').replace(' ', '')
    # Si tiene puntos como separador de miles y coma como decimal → dejarlo como está
    # Si es número puro con puntos como miles → convertir a entero y volver a formatear
    s_clean = s.replace('.', '').replace(',', '.')
    try:
        n = float(s_clean)
        # Formatear como entero si no tiene decimales significativos
        if n == int(n):
            return f"$ {int(n):,}".replace(',', '.')
        return f"$ {n:,.2f}".replace(',', '.')
    except Exception:
        return s

def clean_str(val):
    if val is None:
        return ''
    s = str(val).strip()
    if s.lower() in ('nan', 'none'):
        return ''
    return s

def main():
    excel_path = Path(sys.argv[1]) if len(sys.argv) > 1 else EXCEL_DEFAULT

    if not excel_path.exists():
        print(f"ERROR: No se encontró el archivo:\n  {excel_path}")
        print("\nUsá: python3 migrate_excel.py /ruta/al/archivo.xlsx")
        sys.exit(1)

    print(f"Leyendo: {excel_path.name}")
    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)

    if 'trasladable' not in wb.sheetnames:
        print(f"ERROR: No se encontró la hoja 'trasladable'. Hojas disponibles: {wb.sheetnames}")
        sys.exit(1)

    ws = wb['trasladable']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # La fila 0 es el header
    header = [str(c).strip() if c else '' for c in rows[0]]
    print(f"Columnas detectadas: {header}")

    # Mapeo por posición (basado en el orden conocido del Excel)
    # 0: FECHA INST., 1: FECHA AVAL RECIBIDO, 2: FECHA TASACION RECIBIDA,
    # 3: FECHA CARGA, 4: NUM EXP, 5: TITULAR, 6: PROV, 7: NOMBRE DE PROYECTO,
    # 8: LINEA PROGRAMATICA, 9: MONTO ($), 10: GARANTIA, 11: TEC. DE CARGA,
    # 12: PASO CARGA INICIAL?, 13: SE SOLICITÓ:, 14: TECNICO
    COL = {
        'fecha_instruccion':       0,
        'fecha_aval_recibido':     1,
        'fecha_tasacion_recibida': 2,
        'fecha_carga':             3,
        'num_exp':                 4,
        'titular':                 5,
        'provincia':               6,
        'nombre_proyecto':         7,
        'linea_programatica':      8,
        'monto':                   9,
        'garantia':               10,
        'tec_de_carga':           11,
        'paso_carga_inicial':     12,
        'se_solicito':            13,
        'tecnico':                14,
    }

    # Conectar a DB
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA journal_mode=WAL")

    # Crear tabla si no existe
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS expedientes (
            id                      INTEGER PRIMARY KEY AUTOINCREMENT,
            num_exp                 TEXT NOT NULL,
            fecha_instruccion       TEXT,
            fecha_aval_recibido     TEXT,
            fecha_tasacion_recibida TEXT,
            fecha_carga             TEXT,
            titular                 TEXT,
            provincia               TEXT,
            nombre_proyecto         TEXT,
            linea_programatica      TEXT,
            monto                   TEXT,
            garantia                TEXT,
            tec_de_carga            TEXT,
            paso_carga_inicial      TEXT,
            se_solicito             TEXT,
            tecnico                 TEXT,
            observaciones           TEXT DEFAULT '',
            deleted_at              TEXT,
            created_at              TEXT DEFAULT (datetime('now','localtime')),
            updated_at              TEXT DEFAULT (datetime('now','localtime'))
        );
        CREATE UNIQUE INDEX IF NOT EXISTS idx_num_exp ON expedientes(num_exp);
    """)

    existentes = set(r[0] for r in conn.execute("SELECT num_exp FROM expedientes").fetchall())

    ok = 0; skipped = 0; errors = []

    for i, row in enumerate(rows[1:], start=2):
        if not row or all(c is None for c in row):
            skipped += 1
            continue

        def get(col_idx):
            return row[col_idx] if col_idx < len(row) else None

        num_exp = clean_str(get(COL['num_exp']))
        if not num_exp:
            skipped += 1
            continue

        # Normalizar formato expediente
        num_exp = num_exp.upper().strip()

        if num_exp in existentes:
            skipped += 1
            continue

        try:
            conn.execute("""
                INSERT INTO expedientes
                    (num_exp, fecha_instruccion, fecha_aval_recibido, fecha_tasacion_recibida,
                     fecha_carga, titular, provincia, nombre_proyecto, linea_programatica,
                     monto, garantia, tec_de_carga, paso_carga_inicial, se_solicito, tecnico)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                num_exp,
                fmt_date(get(COL['fecha_instruccion'])),
                fmt_date(get(COL['fecha_aval_recibido'])),
                fmt_date(get(COL['fecha_tasacion_recibida'])),
                fmt_date(get(COL['fecha_carga'])),
                clean_str(get(COL['titular'])),
                clean_str(get(COL['provincia'])),
                clean_str(get(COL['nombre_proyecto'])),
                clean_str(get(COL['linea_programatica'])),
                clean_monto(get(COL['monto'])),
                clean_str(get(COL['garantia'])),
                clean_str(get(COL['tec_de_carga'])),
                clean_str(get(COL['paso_carga_inicial'])),
                clean_str(get(COL['se_solicito'])),
                clean_str(get(COL['tecnico'])),
            ))
            existentes.add(num_exp)
            ok += 1
        except Exception as e:
            errors.append(f"  Fila {i}: {num_exp} → {e}")

        if ok % 200 == 0 and ok > 0:
            conn.commit()
            print(f"  {ok} importados...")

    conn.commit()
    conn.close()

    print(f"\n{'='*50}")
    print(f"  Importados:  {ok}")
    print(f"  Omitidos:    {skipped}")
    print(f"  Errores:     {len(errors)}")
    if errors:
        print("\nErrores:")
        for e in errors[:20]:
            print(e)
    print(f"{'='*50}")
    print(f"\nDB guardada en: {DB_PATH}")
    print("Listo. Ahora iniciá el servidor: python3 server.py")

if __name__ == '__main__':
    main()
